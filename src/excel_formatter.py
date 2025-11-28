"""Excel 서식 적용"""
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class ExcelFormatter:
    """Excel 파일 서식 적용 클래스"""

    def apply_header_format(self, worksheet):
        """헤더 서식 적용 (맑은 고딕 12pt Bold, 배경/글자 색상)

        Args:
            worksheet: openpyxl Worksheet 객체
        """
        # 헤더 행 (1행)
        for cell in worksheet[1]:
            # 폰트: 맑은 고딕, 12pt, Bold
            cell.font = Font(name="맑은 고딕", size=12, bold=True, color="9C5700")

            # 배경색: #FFEB9C
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            # 정렬: 가운데
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def apply_data_format(self, worksheet):
        """데이터 셀 서식 적용 (맑은 고딕 10pt, 테두리)

        Args:
            worksheet: openpyxl Worksheet 객체
        """
        # 테두리 스타일
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        # 모든 셀에 서식 적용
        for row in worksheet.iter_rows(min_row=1):
            for cell in row:
                # 폰트: 맑은 고딕, 10pt
                if cell.row == 1:
                    # 헤더는 이미 apply_header_format에서 처리
                    pass
                else:
                    cell.font = Font(name="맑은 고딕", size=10)

                # 테두리
                cell.border = thin_border

    def freeze_panes(self, worksheet):
        """틀 고정 적용 (A2 - 헤더 행 고정)

        Args:
            worksheet: openpyxl Worksheet 객체
        """
        worksheet.freeze_panes = "A2"
