"""NC/GL 병합 테스트"""
import os
import tempfile
import shutil
import pytest
import pandas as pd
from openpyxl import Workbook


class TestNCGLMerger:
    """NCGLMerger 클래스 테스트"""

    @pytest.fixture
    def temp_dir_with_files(self):
        """테스트용 8개 언어 파일이 있는 임시 디렉토리"""
        temp_dir = tempfile.mkdtemp()

        # 8개 언어 파일 생성 (PRD 4.3)
        language_files = [
            ("StringEnglish.xlsx", "EN"),
            ("StringTraditionalChinese.xlsx", "CT"),
            ("StringSimplifiedChinese.xlsx", "CS"),
            ("StringJapanese.xlsx", "JA"),
            ("StringThai.xlsx", "TH"),
            ("StringSpanish.xlsx", "ES"),
            ("StringPortuguese.xlsx", "PT"),
            ("StringRussian.xlsx", "RU"),
        ]

        # 각 파일의 고정 구조: Table, KEY, Source, Target, Status, NOTE
        for filename, lang_code in language_files:
            file_path = os.path.join(temp_dir, filename)
            wb = Workbook()
            ws = wb.active

            # 헤더
            ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])

            # 데이터 (2개 행)
            ws.append(["TABLE1", "KEY001", "Hello", f"Hello_{lang_code}", "Translated", "Note1"])
            ws.append(["TABLE2", "KEY002", "World", f"World_{lang_code}", "Translated", "Note2"])

            wb.save(file_path)
            wb.close()

        yield temp_dir
        shutil.rmtree(temp_dir, ignore_errors=True)

    def test_merge_ncgl_loads_8_language_files(self, temp_dir_with_files):
        """8개 언어 파일을 읽어야 함"""
        from src.ncgl_merger import NCGLMerger

        merger = NCGLMerger()
        result = merger.merge_ncgl(temp_dir_with_files, "250115", "42")

        assert result["success"] is True, f"병합이 성공해야 합니다. 에러: {result.get('error', 'Unknown')}"
        assert "output_file" in result, "결과에 output_file이 있어야 합니다"

    def test_merge_ncgl_creates_correct_output_structure(self, temp_dir_with_files):
        """NC/GL 병합 결과가 올바른 헤더 구조를 가져야 함"""
        from src.ncgl_merger import NCGLMerger

        merger = NCGLMerger()
        result = merger.merge_ncgl(temp_dir_with_files, "250115", "42")

        # 출력 파일 확인
        output_df = pd.read_excel(result["output_file"])

        # PRD 4.4 헤더 확인
        required_headers = [
            "Key", "Source", "Target_EN", "Target_CT", "Target_CS",
            "Target_JA", "Target_TH", "Target_ES", "Target_PT", "Target_RU",
            "Comment", "TableName", "Status"
        ]
        for header in required_headers:
            assert header in output_df.columns, f"{header} 헤더가 있어야 합니다"

    def test_merge_ncgl_uses_en_as_master(self, temp_dir_with_files):
        """EN 파일을 마스터로 사용해야 함"""
        from src.ncgl_merger import NCGLMerger

        merger = NCGLMerger()
        result = merger.merge_ncgl(temp_dir_with_files, "250115", "42")

        # 출력 파일 확인
        output_df = pd.read_excel(result["output_file"])

        # KEY는 EN 파일 기준
        assert len(output_df) == 2, "EN 파일의 2개 행이 있어야 합니다"
        assert output_df["Key"].iloc[0] == "KEY001", "첫 번째 KEY가 일치해야 합니다"
        assert output_df["Key"].iloc[1] == "KEY002", "두 번째 KEY가 일치해야 합니다"

    def test_merge_ncgl_validates_field_consistency(self):
        """Table, Source, Status, NOTE 불일치 시 에러를 반환해야 함"""
        from src.ncgl_merger import NCGLMerger
        import tempfile

        temp_dir = tempfile.mkdtemp()

        try:
            # EN 파일 생성
            en_path = os.path.join(temp_dir, "StringEnglish.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
            ws.append(["TABLE1", "KEY001", "Hello", "Hello_EN", "Translated", "Note1"])
            wb.save(en_path)
            wb.close()

            # CT 파일 생성 (Source 불일치)
            ct_path = os.path.join(temp_dir, "StringTraditionalChinese.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
            ws.append(["TABLE1", "KEY001", "Different", "Hello_CT", "Translated", "Note1"])  # Source 다름
            wb.save(ct_path)
            wb.close()

            # 나머지 파일들도 생성 (간단히)
            for filename in ["StringSimplifiedChinese.xlsx", "StringJapanese.xlsx",
                           "StringThai.xlsx", "StringSpanish.xlsx",
                           "StringPortuguese.xlsx", "StringRussian.xlsx"]:
                file_path = os.path.join(temp_dir, filename)
                wb = Workbook()
                ws = wb.active
                ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
                ws.append(["TABLE1", "KEY001", "Hello", "Hello_X", "Translated", "Note1"])
                wb.save(file_path)
                wb.close()

            merger = NCGLMerger()
            result = merger.merge_ncgl(temp_dir, "250115", "42")

            # PRD 4.6 검증 규칙: 불일치 발견 시 작업 중단
            assert result["success"] is False, "필드 불일치 시 실패해야 합니다"
            assert "Source" in result["error"] or "다릅니다" in result["error"], "불일치 에러 메시지가 있어야 합니다"

        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_merge_ncgl_generates_correct_filename(self, temp_dir_with_files):
        """출력 파일명이 {YYMMDD}_M{milestone}_StringALL.xlsx 형식이어야 함"""
        from src.ncgl_merger import NCGLMerger

        merger = NCGLMerger()
        result = merger.merge_ncgl(temp_dir_with_files, "250115", "42")

        # PRD 4.4 파일명 검증
        output_filename = os.path.basename(result["output_file"])
        assert output_filename == "250115_M42_StringALL.xlsx", f"파일명이 일치해야 합니다: {output_filename}"

    def test_merge_ncgl_handles_nan_values(self, temp_dir_with_files):
        """NaN/inf 값을 빈 문자열로 변환해야 함"""
        from src.ncgl_merger import NCGLMerger
        import tempfile
        import math

        temp_dir = tempfile.mkdtemp()

        try:
            # NaN 값이 있는 파일 생성
            for filename in ["StringEnglish.xlsx", "StringTraditionalChinese.xlsx",
                           "StringSimplifiedChinese.xlsx", "StringJapanese.xlsx",
                           "StringThai.xlsx", "StringSpanish.xlsx",
                           "StringPortuguese.xlsx", "StringRussian.xlsx"]:
                file_path = os.path.join(temp_dir, filename)
                wb = Workbook()
                ws = wb.active
                ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
                ws.append(["TABLE1", "KEY001", "Hello", None, "Translated", None])  # NaN 값
                wb.save(file_path)
                wb.close()

            merger = NCGLMerger()
            result = merger.merge_ncgl(temp_dir, "250115", "42")

            # 출력 파일 확인 (빈 문자열을 NaN으로 읽지 않도록)
            output_df = pd.read_excel(result["output_file"], keep_default_na=False)

            # NaN 값이 빈 문자열로 변환되어야 함
            for col in output_df.columns:
                for val in output_df[col]:
                    if pd.isna(val):
                        assert False, f"{col} 열에 NaN 값이 있으면 안 됩니다"

        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
