"""M4/GL 통합 병합 테스트"""
import os
import tempfile
import shutil
import pytest
import pandas as pd
from openpyxl import Workbook


class TestM4GLMerger:
    """M4GLMerger 클래스 테스트 - DIALOGUE + STRING 통합 병합"""

    @pytest.fixture
    def temp_dir_with_all_files(self):
        """테스트용 DIALOGUE + STRING 파일이 있는 임시 디렉토리"""
        temp_dir = tempfile.mkdtemp()

        # CINEMATIC_DIALOGUE.xlsm 생성
        cinematic_path = os.path.join(temp_dir, "CINEMATIC_DIALOGUE.xlsm")
        wb = Workbook()
        wb.create_sheet("Sheet2")
        ws = wb["Sheet2"]
        ws.append([None] * 30)
        ws.append([""] * 7 + ["String ID", "NPC ID"] + [""] * 2 + ["KO (M)", "KO (F)", "EN (M)", "EN (F)"] + [""] * 14)
        for _ in range(7):
            ws.append([None] * 30)
        ws.append([""] * 7 + ["1", "100"] + [""] * 2 + ["안녕", "안녕하세요", "Hello", "Hi"] + [""] * 14)
        wb.save(cinematic_path)
        wb.close()

        # SMALLTALK_DIALOGUE.xlsm 생성
        smalltalk_path = os.path.join(temp_dir, "SMALLTALK_DIALOGUE.xlsm")
        wb = Workbook()
        wb.create_sheet("Sheet2")
        ws = wb["Sheet2"]
        ws.append([None] * 31)
        ws.append([""] * 7 + ["String ID", "NPC ID"] + [""] * 3 + ["KO (M)", "KO (F)", "EN (M)", "EN (F)"] + [""] * 14)
        for _ in range(2):
            ws.append([None] * 31)
        ws.append([""] * 7 + ["2", "101"] + [""] * 3 + ["감사", "감사합니다", "Thanks", "Thank you"] + [""] * 14)
        wb.save(smalltalk_path)
        wb.close()

        # NPC.xlsm 생성
        npc_path = os.path.join(temp_dir, "NPC.xlsm")
        wb = Workbook()
        ws = wb.active
        ws.title = "NPC"
        ws.append([None] * 10)
        ws.append([""] * 7 + ["NPC ID"] + [""] + ["Speaker Name"])
        ws.append([""] * 7 + ["100"] + [""] + ["캐릭터A"])
        ws.append([""] * 7 + ["101"] + [""] + ["캐릭터B"])
        wb.save(npc_path)
        wb.close()

        # STRING 파일 8개 생성 (필수 검증 통과용)
        string_files = [
            ("SEQUENCE_DIALOGUE.xlsm", 7, None, 10, 11),
            ("STRING_BUILTIN.xlsm", 7, 21, 8, 9),
            ("STRING_MAIL.xlsm", 7, None, 8, 9),
            ("STRING_MESSAGE.xlsm", 7, 21, 8, 9),
            ("STRING_NPC.xlsm", 7, 20, 9, 10),
            ("STRING_QUESTTEMPLATE.xlsm", 7, 0, 12, 13),
            ("STRING_TEMPLATE.xlsm", 7, 19, 8, 9),
            ("STRING_TOOLTIP.xlsm", 7, 8, 11, 12),
        ]

        for filename, str_id_col, note_col, ko_col, en_col in string_files:
            string_path = os.path.join(temp_dir, filename)
            wb = Workbook()
            ws = wb.active
            ws.append([None] * 25)
            header = [""] * 25
            header[str_id_col] = "String ID"
            if note_col is not None:
                header[note_col] = "NOTE"
            header[ko_col] = "KO"
            header[en_col] = "EN"
            ws.append(header)
            for _ in range(1):
                ws.append([None] * 25)
            data_row = [""] * 25
            data_row[str_id_col] = "2001"
            if note_col is not None:
                data_row[note_col] = "테스트"
            data_row[ko_col] = "문자열"
            data_row[en_col] = "String"
            ws.append(data_row)
            wb.save(string_path)
            wb.close()

        yield temp_dir
        shutil.rmtree(temp_dir, ignore_errors=True)

    def test_merge_all_executes_dialogue_then_string(self, temp_dir_with_all_files):
        """DIALOGUE 병합 → STRING 병합 순차 실행"""
        from src.m4gl_merger import M4GLMerger

        merger = M4GLMerger()
        result = merger.merge_all(temp_dir_with_all_files)

        assert result["success"] is True, f"통합 병합이 성공해야 합니다. 에러: {result.get('error', 'Unknown')}"
        assert "dialogue_result" in result, "DIALOGUE 병합 결과가 있어야 합니다"
        assert "string_result" in result, "STRING 병합 결과가 있어야 합니다"

    def test_merge_all_creates_two_output_files(self, temp_dir_with_all_files):
        """통합 병합은 2개 파일을 생성해야 함"""
        from src.m4gl_merger import M4GLMerger

        merger = M4GLMerger()
        result = merger.merge_all(temp_dir_with_all_files)

        # 출력 파일 확인
        assert os.path.exists(result["dialogue_result"]["output_file"]), "DIALOGUE 파일이 생성되어야 합니다"
        assert os.path.exists(result["string_result"]["output_file"]), "STRING 파일이 생성되어야 합니다"

        # 파일명 확인
        dialogue_filename = os.path.basename(result["dialogue_result"]["output_file"])
        string_filename = os.path.basename(result["string_result"]["output_file"])

        assert "MIR4_MASTER_DIALOGUE" in dialogue_filename, "DIALOGUE 파일명이 올바라야 합니다"
        assert "MIR4_MASTER_STRING" in string_filename, "STRING 파일명이 올바라야 합니다"

    def test_merge_all_returns_total_count(self, temp_dir_with_all_files):
        """통합 병합은 총 행 수를 반환해야 함"""
        from src.m4gl_merger import M4GLMerger

        merger = M4GLMerger()
        result = merger.merge_all(temp_dir_with_all_files)

        assert "total_rows" in result, "총 행 수가 있어야 합니다"
        assert result["total_rows"] >= 3, "최소 3개 행이 있어야 합니다 (DIALOGUE 2 + STRING 1)"
