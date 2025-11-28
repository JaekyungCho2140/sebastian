"""M4/GL DIALOGUE 병합 테스트"""
import os
import tempfile
import shutil
import pytest
import pandas as pd
from openpyxl import Workbook


class TestDialogueMerger:
    """DialogueMerger 클래스 테스트"""

    @pytest.fixture
    def temp_dir_with_files(self):
        """테스트용 Excel 파일이 있는 임시 디렉토리"""
        temp_dir = tempfile.mkdtemp()

        # CINEMATIC_DIALOGUE.xlsm 생성 (간단한 구조)
        cinematic_path = os.path.join(temp_dir, "CINEMATIC_DIALOGUE.xlsm")
        wb = Workbook()
        # Sheet1 생성 후 Sheet2 추가
        wb.create_sheet("Sheet2")
        ws = wb["Sheet2"]
        # 헤더 (2행)
        ws.append([None] * 30)
        ws.append([""] * 7 + ["String ID", "NPC ID"] + [""] * 2 + ["KO (M)", "KO (F)", "EN (M)", "EN (F)"] + [""] * 14)
        # 빈 행들
        for _ in range(7):
            ws.append([None] * 30)
        # 데이터 (10행부터)
        ws.append([""] * 7 + ["1", "100"] + [""] * 2 + ["안녕", "안녕하세요", "Hello", "Hi"] + [""] * 14)
        wb.save(cinematic_path)
        wb.close()

        # SMALLTALK_DIALOGUE.xlsm 생성
        smalltalk_path = os.path.join(temp_dir, "SMALLTALK_DIALOGUE.xlsm")
        wb = Workbook()
        wb.create_sheet("Sheet2")
        ws = wb["Sheet2"]
        ws.append([None] * 31)
        ws.append([""] * 7 + ["String ID", "NPC ID"] + [""] * 3 + ["KO (M)", "KO (F)", "EN (M)", "EN (F)"] + [""] * 14)
        for _ in range(2):
            ws.append([None] * 31)
        ws.append([""] * 7 + ["2", "101"] + [""] * 3 + ["감사", "감사합니다", "Thanks", "Thank you"] + [""] * 14)
        wb.save(smalltalk_path)
        wb.close()

        # NPC.xlsm 생성
        npc_path = os.path.join(temp_dir, "NPC.xlsm")
        wb = Workbook()
        ws = wb.active
        ws.title = "NPC"
        ws.append([None] * 10)
        ws.append([""] * 7 + ["NPC ID"] + [""] + ["Speaker Name"])
        ws.append([""] * 7 + ["100"] + [""] + ["캐릭터A"])
        ws.append([""] * 7 + ["101"] + [""] + ["캐릭터B"])
        wb.save(npc_path)
        wb.close()

        yield temp_dir
        shutil.rmtree(temp_dir, ignore_errors=True)

    def test_merge_dialogue_combines_cinematic_and_smalltalk(self, temp_dir_with_files):
        """CINEMATIC과 SMALLTALK 데이터를 병합해야 함"""
        from src.dialogue_merger import DialogueMerger

        merger = DialogueMerger()

        result = merger.merge_dialogue(temp_dir_with_files)

        # 에러 확인
        if not result["success"]:
            print(f"병합 실패: {result.get('error', 'Unknown error')}")

        assert result["success"] is True, f"병합이 성공해야 합니다. 에러: {result.get('error', 'Unknown')}"
        assert "output_file" in result, "결과에 output_file이 있어야 합니다"
        assert result["row_count"] >= 2, "최소 2개 행이 병합되어야 합니다"

    def test_merge_dialogue_filters_empty_en_m_rows(self, temp_dir_with_files):
        """EN (M)이 비어있는 행은 제거되어야 함"""
        from src.dialogue_merger import DialogueMerger

        merger = DialogueMerger()
        result = merger.merge_dialogue(temp_dir_with_files)

        # 출력 파일 확인
        output_df = pd.read_excel(result["output_file"])

        # EN (M) 열이 모두 비어있지 않아야 함
        en_m_col = output_df["EN (M)"]
        assert not en_m_col.isna().any(), "EN (M)이 빈 행이 있으면 안 됩니다"
