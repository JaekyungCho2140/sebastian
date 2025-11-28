"""LY/GL 병합 테스트"""
import os
import tempfile
import shutil
import pytest
import pandas as pd
from openpyxl import Workbook


class TestLYGLMerger:
    """LYGLMerger 클래스 테스트"""

    @pytest.fixture
    def temp_dir_with_files(self):
        """테스트용 7개 언어 파일이 있는 임시 디렉토리"""
        temp_dir = tempfile.mkdtemp()

        # PRD 5.1.2: 7개 언어 파일 생성
        language_files = [
            "251104_EN.xlsx",
            "251104_CT.xlsx",
            "251104_CS.xlsx",
            "251104_JA.xlsx",
            "251104_TH.xlsx",
            "251104_PT-BR.xlsx",
            "251104_RU.xlsx",
        ]

        # PRD 5.1.4: 각 파일의 고정 구조 (Table, KEY, Source, Target, Status, NOTE)
        for filename in language_files:
            file_path = os.path.join(temp_dir, filename)
            wb = Workbook()
            ws = wb.active

            # 헤더
            ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])

            # 데이터 (2개 행)
            lang_code = filename.split("_")[1].replace(".xlsx", "")
            ws.append(["TABLE1", "KEY001", "Hello", f"Hello_{lang_code}", "Translated", "Note1"])
            ws.append(["TABLE2", "KEY002", "World", f"World_{lang_code}", "Translated", "Note2"])

            wb.save(file_path)
            wb.close()

        yield temp_dir
        shutil.rmtree(temp_dir, ignore_errors=True)

    def test_merge_lygl_loads_7_language_files(self, temp_dir_with_files):
        """7개 언어 파일을 읽어야 함"""
        from src.lygl_merger import LYGLMerger

        merger = LYGLMerger()
        result = merger.merge_lygl(temp_dir_with_files)

        assert result["success"] is True, f"병합이 성공해야 합니다. 에러: {result.get('error', 'Unknown')}"
        assert "output_file" in result, "결과에 output_file이 있어야 합니다"

    def test_merge_lygl_creates_correct_output_structure(self, temp_dir_with_files):
        """LY/GL 병합 결과가 올바른 헤더 구조를 가져야 함"""
        from src.lygl_merger import LYGLMerger

        merger = LYGLMerger()
        result = merger.merge_lygl(temp_dir_with_files)

        # 출력 파일 확인
        output_df = pd.read_excel(result["output_file"])

        # PRD 5.1.3 헤더 확인
        required_headers = [
            "Table", "KEY", "Source", "Target_EN", "Target_CT", "Target_CS",
            "Target_JA", "Target_TH", "Target_PT-BR", "Target_RU", "Status", "NOTE"
        ]
        for header in required_headers:
            assert header in output_df.columns, f"{header} 헤더가 있어야 합니다"

    def test_merge_lygl_validates_exactly_7_files(self):
        """정확히 7개 파일이 필요함 (PRD 5.1.5)"""
        from src.lygl_merger import LYGLMerger
        import tempfile

        temp_dir = tempfile.mkdtemp()

        try:
            # 6개 파일만 생성 (RU 누락)
            for lang in ["EN", "CT", "CS", "JA", "TH", "PT-BR"]:
                file_path = os.path.join(temp_dir, f"251104_{lang}.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
                ws.append(["TABLE1", "KEY001", "Hello", "Hello", "Translated", "Note"])
                wb.save(file_path)
                wb.close()

            merger = LYGLMerger()
            result = merger.merge_lygl(temp_dir)

            # PRD 5.1.6: 파일 수 불일치 에러
            assert result["success"] is False, "파일 수가 맞지 않으면 실패해야 합니다"
            assert "7개 언어 파일이 필요합니다" in result["error"], "파일 수 에러 메시지가 있어야 합니다"

        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_merge_lygl_validates_key_consistency(self):
        """KEY 일치 검증 (PRD 5.1.5)"""
        from src.lygl_merger import LYGLMerger
        import tempfile

        temp_dir = tempfile.mkdtemp()

        try:
            # EN 파일 생성
            en_path = os.path.join(temp_dir, "251104_EN.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
            ws.append(["TABLE1", "KEY001", "Hello", "Hello_EN", "Translated", "Note"])
            wb.save(en_path)
            wb.close()

            # CT 파일 생성 (KEY 다름)
            ct_path = os.path.join(temp_dir, "251104_CT.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
            ws.append(["TABLE1", "KEY999", "Hello", "Hello_CT", "Translated", "Note"])  # 다른 KEY
            wb.save(ct_path)
            wb.close()

            # 나머지 파일들 생성
            for lang in ["CS", "JA", "TH", "PT-BR", "RU"]:
                file_path = os.path.join(temp_dir, f"251104_{lang}.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
                ws.append(["TABLE1", "KEY001", "Hello", "Hello", "Translated", "Note"])
                wb.save(file_path)
                wb.close()

            merger = LYGLMerger()
            result = merger.merge_lygl(temp_dir)

            # PRD 5.1.6: KEY 불일치 에러
            assert result["success"] is False, "KEY 불일치 시 실패해야 합니다"
            assert "KEY" in result["error"] or "없습니다" in result["error"], "KEY 불일치 에러 메시지가 있어야 합니다"

        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_merge_lygl_validates_field_consistency(self):
        """Table, Source, Status, NOTE 일치 검증 (PRD 5.1.5)"""
        from src.lygl_merger import LYGLMerger
        import tempfile

        temp_dir = tempfile.mkdtemp()

        try:
            # EN 파일 생성
            en_path = os.path.join(temp_dir, "251104_EN.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
            ws.append(["TABLE1", "KEY001", "Hello", "Hello_EN", "Translated", "Note1"])
            wb.save(en_path)
            wb.close()

            # CT 파일 생성 (Table 다름)
            ct_path = os.path.join(temp_dir, "251104_CT.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
            ws.append(["TABLE_DIFFERENT", "KEY001", "Hello", "Hello_CT", "Translated", "Note1"])  # Table 다름
            wb.save(ct_path)
            wb.close()

            # 나머지 파일들 생성
            for lang in ["CS", "JA", "TH", "PT-BR", "RU"]:
                file_path = os.path.join(temp_dir, f"251104_{lang}.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE"])
                ws.append(["TABLE1", "KEY001", "Hello", "Hello", "Translated", "Note1"])
                wb.save(file_path)
                wb.close()

            merger = LYGLMerger()
            result = merger.merge_lygl(temp_dir)

            # PRD 5.1.6: 데이터 불일치 에러
            assert result["success"] is False, "필드 불일치 시 실패해야 합니다"
            assert "Table" in result["error"] or "다릅니다" in result["error"], "필드 불일치 에러 메시지가 있어야 합니다"

        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_merge_lygl_auto_detects_date_prefix(self, temp_dir_with_files):
        """파일명에서 날짜 접두사를 자동 감지해야 함"""
        from src.lygl_merger import LYGLMerger

        merger = LYGLMerger()
        result = merger.merge_lygl(temp_dir_with_files)

        # 출력 파일명이 251104로 시작해야 함
        output_filename = os.path.basename(result["output_file"])
        assert output_filename.startswith("251104_"), f"파일명이 251104_로 시작해야 합니다: {output_filename}"
