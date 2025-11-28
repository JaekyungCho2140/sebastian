"""Excel 서식 적용 테스트"""
import os
import tempfile
import shutil
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class TestExcelFormatter:
    """ExcelFormatter 클래스 테스트"""

    @pytest.fixture
    def temp_excel_file(self):
        """임시 Excel 파일 생성"""
        temp_dir = tempfile.mkdtemp()
        excel_path = os.path.join(temp_dir, "test.xlsx")

        # 간단한 Excel 파일 생성
        wb = Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2", "Header3"])
        ws.append(["Data1", "Data2", "Data3"])
        wb.save(excel_path)
        wb.close()

        yield excel_path
        shutil.rmtree(temp_dir, ignore_errors=True)

    def test_apply_header_format_sets_font_and_color(self, temp_excel_file):
        """헤더 서식을 적용해야 함 (맑은 고딕 12pt Bold, 배경/글자 색상)"""
        from src.excel_formatter import ExcelFormatter
        from openpyxl import load_workbook

        formatter = ExcelFormatter()

        # Excel 파일 로드
        wb = load_workbook(temp_excel_file)
        ws = wb.active

        # 헤더 서식 적용
        formatter.apply_header_format(ws)

        # 헤더 행(1행) 서식 확인
        header_cell = ws['A1']
        assert header_cell.font.name == "맑은 고딕", "헤더 폰트가 맑은 고딕이어야 합니다"
        assert header_cell.font.size == 12, "헤더 폰트 크기가 12pt여야 합니다"
        assert header_cell.font.bold is True, "헤더가 Bold여야 합니다"
        # openpyxl은 ARGB 형식 사용 (00FFEB9C)
        assert "FFEB9C" in header_cell.fill.start_color.rgb, "헤더 배경색이 #FFEB9C여야 합니다"
        assert "9C5700" in header_cell.font.color.rgb, "헤더 글자색이 #9C5700이어야 합니다"

        wb.close()

    def test_apply_data_format_sets_font_and_border(self, temp_excel_file):
        """데이터 셀 서식을 적용해야 함 (맑은 고딕 10pt, 테두리)"""
        from src.excel_formatter import ExcelFormatter
        from openpyxl import load_workbook

        formatter = ExcelFormatter()

        wb = load_workbook(temp_excel_file)
        ws = wb.active

        # 데이터 서식 적용
        formatter.apply_data_format(ws)

        # 데이터 행(2행) 서식 확인
        data_cell = ws['A2']
        assert data_cell.font.name == "맑은 고딕", "데이터 폰트가 맑은 고딕이어야 합니다"
        assert data_cell.font.size == 10, "데이터 폰트 크기가 10pt여야 합니다"
        assert data_cell.border.top.style == "thin", "테두리가 thin이어야 합니다"

        wb.close()

    def test_freeze_panes_freezes_at_a2(self, temp_excel_file):
        """틀 고정을 A2에 적용해야 함 (헤더 행 고정)"""
        from src.excel_formatter import ExcelFormatter
        from openpyxl import load_workbook

        formatter = ExcelFormatter()

        wb = load_workbook(temp_excel_file)
        ws = wb.active

        # 틀 고정 적용
        formatter.freeze_panes(ws)

        # 틀 고정 확인
        assert ws.freeze_panes == "A2", "틀 고정이 A2에 설정되어야 합니다"

        wb.close()
